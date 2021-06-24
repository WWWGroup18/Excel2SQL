import openpyxl
from pathlib import Path
import datetime

def quote_fixer(x): #this function is used to replace quotes to make it SQL friendly
    ont= x.replace('"', '""')
    twt= ont.replace("'", "''")
    return twt

st="INSERT INTO products(p_category, p_name, p_image, p_description,p_specs, p_amazon, p_newegg, p_bestbuy) VALUES ("
category=101
# Setting the path to the xlsx file:
xlsx_file = Path('input', 'eighteentech data.xlsx')

wb_obj = openpyxl.load_workbook(xlsx_file) # opening xlsx file
opname="./output/"+str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))+".txt" #unique output file name
f = open(opname, "wb") #output file
sheets= wb_obj.sheetnames #getting names of sheets in order

for sheetname in sheets: #iterate through sheets
    ws= wb_obj[sheetname]
    num=2 #row number
    while ws["A"+str(num)].value is not None :
        name= quote_fixer(ws["A"+str(num)].value)
        image_link= ws["B"+str(num)].value
        description= quote_fixer(ws["C"+str(num)].value)
        specs= quote_fixer(ws["D"+str(num)].value)

        if ws["E"+str(num)].value is None:
            amazon=""
        else:
            amazon= ws["E"+str(num)].value
        if ws["F"+str(num)].value is None:
            newegg=""
        else:
            newegg= ws["F"+str(num)].value
        if ws["G"+str(num)].value is None:
            bestbuy=""
        else:
            bestbuy= ws["G"+str(num)].value

        op=st+"\""+str(category)+"\", \""+name+"\", \""+image_link+"\", \""+description+"\", \""+ specs+ "\", \""+ amazon+"\", \""+newegg+"\", \""+bestbuy+"\");\n"
        f.write(op.encode("utf-8")) #write to output file
        num=num+1
    category=category+1
f.close()
